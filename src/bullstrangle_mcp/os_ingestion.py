from __future__ import annotations

import json
from datetime import date
from hashlib import sha256
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .database import DEFAULT_DB_PATH, connect, initialize_database


HEADER_ROW = 9
DATA_START_ROW = 10
LIVE_VALUE_COLUMNS = [
    "live_stock_price",
    "live_stock_iv",
    "sell_call_strike",
    "sell_call_bid",
    "sell_put_strike",
    "sell_put_bid",
    "buy_put_strike",
    "buy_put_ask",
    "total_credit",
]


def ingest_os_workbook(
    workbook_path: str | Path,
    db_path: str | Path = DEFAULT_DB_PATH,
    trading_date: str | None = None,
    regenerate_if_stale: bool = False,
) -> dict[str, Any]:
    initialize_database(db_path)
    path = Path(workbook_path)
    if not path.exists():
        raise FileNotFoundError(path)

    formula_wb = load_workbook(path, data_only=False, read_only=True)
    values_wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if "OS_Live" not in values_wb.sheetnames:
            raise ValueError("Workbook is missing required sheet: OS_Live")
        formula_ws = formula_wb["OS_Live"]
        values_ws = values_wb["OS_Live"]

        headers = _read_headers(values_ws)
        records = _read_records(values_ws, headers)
        formula_cell_count = _count_formula_cells(formula_ws, headers, len(records))
        metadata = _read_metadata(values_wb)
    finally:
        formula_wb.close()
        values_wb.close()

    if not records:
        raise ValueError("No OS_Live data rows found")

    newsletter_id = int(records[0]["newsletter_id"])
    newsletter_date = str(records[0]["newsletter_date"])
    expiration_date = records[0].get("expiration_date")
    workbook_id = _find_workbook_id(db_path, newsletter_id, path, metadata, records[0])
    populated_live_value_count = _count_populated_live_values(records)
    validation = {
        "missing_live_value_count": _count_missing_live_values(records),
        "headers": headers,
    }
    status = "ingested" if populated_live_value_count else "ingested_no_cached_values"

    with connect(db_path) as conn:
        newsletter = conn.execute(
            """
            SELECT newsletter_id, publication_date, target_expiration
            FROM newsletters
            WHERE newsletter_id = ?
            """,
            (newsletter_id,),
        ).fetchone()
        if not newsletter:
            # newsletter_id mismatch — workbook was generated from a different DB state.
            # Look up whether a newsletter for the same date exists under a different ID.
            current = conn.execute(
                "SELECT newsletter_id FROM newsletters WHERE publication_date = ?",
                (newsletter_date,),
            ).fetchone()
            hint = (
                f"  A newsletter for {newsletter_date} exists in this DB as "
                f"newsletter_id={current['newsletter_id']}.\n"
                f"  Re-generate: bullstrangle --db <db> generate-os-workbook "
                f"{newsletter_date} --output-dir outputs/workbooks"
                if current
                else f"  No newsletter for {newsletter_date} found in this DB either — "
                f"ingest the PDF first: bullstrangle --db <db> ingest-pdf <file.pdf>"
            )
            if regenerate_if_stale and current:
                # Auto-regenerate the workbook and re-run ingest against the fresh file.
                from .os_workbooks import generate_os_workbook

                upload_dir = Path(workbook_path).parent
                regen = generate_os_workbook(newsletter_date, db_path, upload_dir)
                return ingest_os_workbook(
                    regen["generated_path"],
                    db_path=db_path,
                    trading_date=trading_date,
                    regenerate_if_stale=False,
                )
            raise ValueError(
                f"Workbook references newsletter_id={newsletter_id} (date={newsletter_date}) "
                f"which does not exist in this database — the workbook was generated from a "
                f"different DB state and is stale.\n{hint}"
            )
        if newsletter_date != newsletter["publication_date"]:
            raise ValueError(
                "Workbook newsletter_date does not match database newsletter: "
                f"{newsletter_date} != {newsletter['publication_date']}"
            )
        if expiration_date and expiration_date != newsletter["target_expiration"]:
            raise ValueError(
                "Workbook expiration_date does not match database newsletter: "
                f"{expiration_date} != {newsletter['target_expiration']}"
            )
        baseline_rows = {
            int(row["entry_id"]): dict(row)
            for row in conn.execute(
                "SELECT * FROM watchlist_entries WHERE newsletter_id = ?", (newsletter_id,)
            ).fetchall()
        }
        _validate_records(records, newsletter_id, newsletter_date, expiration_date, baseline_rows)

        cur = conn.execute(
            """
            INSERT INTO os_evaluation_runs
            (workbook_id, newsletter_id, newsletter_date, expiration_date, trading_date,
             uploaded_path, market_data_as_of, row_count, populated_live_value_count,
             formula_cell_count, status, raw_workbook_hash, validation_json)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                workbook_id,
                newsletter_id,
                newsletter_date,
                expiration_date,
                trading_date or date.today().isoformat(),
                str(path.resolve()),
                trading_date or date.today().isoformat(),
                len(records),
                populated_live_value_count,
                formula_cell_count,
                status,
                _sha256_file(path),
                json.dumps(validation, sort_keys=True),
            ),
        )
        run_id = int(cur.lastrowid)
        for record in records:
            _insert_evaluation_row(conn, run_id, record)
            baseline = baseline_rows.get(int(record["watchlist_entry_id"]))
            if baseline:
                _insert_deviation_row(conn, run_id, record, baseline)
        conn.commit()

    return {
        "run_id": run_id,
        "workbook_id": workbook_id,
        "newsletter_id": newsletter_id,
        "newsletter_date": newsletter_date,
        "expiration_date": expiration_date,
        "trading_date": trading_date or date.today().isoformat(),
        "uploaded_path": str(path.resolve()),
        "row_count": len(records),
        "populated_live_value_count": populated_live_value_count,
        "formula_cell_count": formula_cell_count,
        "status": status,
        "validation": validation,
    }


def _read_headers(ws) -> list[str]:
    headers = []
    for cell in ws[HEADER_ROW]:
        if cell.value is None:
            break
        headers.append(str(cell.value))
    required = {"newsletter_id", "newsletter_date", "watchlist_entry_id", "symbol"}
    missing = required - set(headers)
    if missing:
        raise ValueError(f"OS_Live is missing required columns: {sorted(missing)}")
    return headers


def _read_records(ws, headers: list[str]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for row in ws.iter_rows(
        min_row=DATA_START_ROW,
        max_row=ws.max_row,
        min_col=1,
        max_col=len(headers),
        values_only=True,
    ):
        record = dict(zip(headers, row))
        if not record.get("symbol"):
            continue
        records.append(record)
    return records


def _read_metadata(wb) -> dict[str, Any]:
    if "Metadata" not in wb.sheetnames:
        return {}
    ws = wb["Metadata"]
    metadata: dict[str, Any] = {}
    for key, value in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if key:
            metadata[str(key)] = value
    return metadata


def _count_formula_cells(ws, headers: list[str], row_count: int) -> int:
    total = 0
    for row in ws.iter_rows(
        min_row=DATA_START_ROW,
        max_row=DATA_START_ROW + row_count - 1,
        min_col=1,
        max_col=len(headers),
        values_only=False,
    ):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                total += 1
    return total


def _find_workbook_id(
    db_path: str | Path,
    newsletter_id: int,
    path: Path,
    metadata: dict[str, Any] | None = None,
    first_record: dict[str, Any] | None = None,
) -> int | None:
    resolved = str(path.resolve())
    with connect(db_path) as conn:
        metadata = metadata or {}
        workbook_id = _int_or_none(metadata.get("workbook_id"))
        if workbook_id is not None:
            row = conn.execute(
                """
                SELECT workbook_id
                FROM os_workbooks
                WHERE workbook_id = ? AND newsletter_id = ?
                """,
                (workbook_id, newsletter_id),
            ).fetchone()
            if row:
                return int(row["workbook_id"])

        row = conn.execute(
            """
            SELECT workbook_id
            FROM os_workbooks
            WHERE newsletter_id = ?
              AND (generated_path = ? OR generated_path IS NULL)
            ORDER BY generated_path IS NULL, workbook_id DESC
            LIMIT 1
            """,
            (newsletter_id, resolved),
        ).fetchone()
        if row:
            return int(row["workbook_id"])

        template_version = metadata.get("template_version")
        selector_source = (
            (first_record or {}).get("selector_source")
            or metadata.get("selector_source")
        )
        if template_version and selector_source:
            row = conn.execute(
                """
                SELECT workbook_id
                FROM os_workbooks
                WHERE newsletter_id = ?
                  AND template_version = ?
                  AND selector_source = ?
                ORDER BY workbook_id DESC
                LIMIT 1
                """,
                (newsletter_id, template_version, selector_source),
            ).fetchone()
            if row:
                return int(row["workbook_id"])
    return None


def _validate_records(
    records: list[dict[str, Any]],
    newsletter_id: int,
    newsletter_date: str,
    expiration_date: Any,
    baseline_rows: dict[int, dict[str, Any]],
) -> None:
    seen_entry_ids: set[int] = set()
    for index, record in enumerate(records, start=DATA_START_ROW):
        row_newsletter_id = _int_or_none(record.get("newsletter_id"))
        if row_newsletter_id != newsletter_id:
            raise ValueError(
                f"OS_Live row {index} has mixed newsletter_id: {row_newsletter_id} != {newsletter_id}"
            )
        if str(record.get("newsletter_date")) != newsletter_date:
            raise ValueError(
                f"OS_Live row {index} has mixed newsletter_date: {record.get('newsletter_date')} != {newsletter_date}"
            )
        if expiration_date and record.get("expiration_date") != expiration_date:
            raise ValueError(
                f"OS_Live row {index} has mixed expiration_date: {record.get('expiration_date')} != {expiration_date}"
            )
        entry_id = _int_or_none(record.get("watchlist_entry_id"))
        if entry_id is None:
            raise ValueError(f"OS_Live row {index} is missing watchlist_entry_id")
        if entry_id in seen_entry_ids:
            raise ValueError(f"OS_Live row {index} duplicates watchlist_entry_id: {entry_id}")
        seen_entry_ids.add(entry_id)

        baseline = baseline_rows.get(entry_id)
        if not baseline:
            raise ValueError(f"OS_Live row {index} references unknown watchlist_entry_id: {entry_id}")
        if record.get("symbol") != baseline["symbol"]:
            raise ValueError(
                f"OS_Live row {index} symbol does not match watchlist_entry_id {entry_id}: "
                f"{record.get('symbol')} != {baseline['symbol']}"
            )


def _count_populated_live_values(records: list[dict[str, Any]]) -> int:
    return sum(1 for record in records for col in LIVE_VALUE_COLUMNS if _num(record.get(col)) is not None)


def _count_missing_live_values(records: list[dict[str, Any]]) -> int:
    return sum(1 for record in records for col in LIVE_VALUE_COLUMNS if _num(record.get(col)) is None)


def _insert_evaluation_row(conn, run_id: int, record: dict[str, Any]) -> None:
    conn.execute(
        """
        INSERT INTO os_evaluation_rows
        (run_id, newsletter_id, newsletter_date, expiration_date, watchlist_entry_id, symbol,
         live_stock_price, live_stock_iv, live_sector, live_industry, live_earnings_date,
         perf_1m, perf_3m, sma_50d, sma_200d, iv_rv_percentile, atr_percent, short_ratio,
         sell_call_strike, sell_call_bid, sell_call_delta, sell_put_strike, sell_put_bid,
         sell_put_delta, buy_put_strike, buy_put_ask, buy_put_delta, total_credit,
         bull_strangle_return_pct, call_distance_pct, put_distance_pct, buy_put_distance_pct,
         call_prob_otm, put_prob_otm, prob_both_otm, selector_source, call_selector_pct,
         put_selector_pct, buy_put_selector_pct, raw_row_json)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            run_id,
            record.get("newsletter_id"),
            record.get("newsletter_date"),
            record.get("expiration_date"),
            record.get("watchlist_entry_id"),
            record.get("symbol"),
            _num(record.get("live_stock_price")),
            _num(record.get("live_stock_iv")),
            record.get("live_sector"),
            record.get("live_industry"),
            _string_or_none(record.get("live_earnings_date")),
            _num(record.get("perf_1m")),
            _num(record.get("perf_3m")),
            _num(record.get("sma_50d")),
            _num(record.get("sma_200d")),
            _num(record.get("iv_rv_percentile")),
            _num(record.get("atr_percent")),
            _num(record.get("short_ratio")),
            _num(record.get("sell_call_strike")),
            _num(record.get("sell_call_bid")),
            _num(record.get("sell_call_delta")),
            _num(record.get("sell_put_strike")),
            _num(record.get("sell_put_bid")),
            _num(record.get("sell_put_delta")),
            _num(record.get("buy_put_strike")),
            _num(record.get("buy_put_ask")),
            _num(record.get("buy_put_delta")),
            _num(record.get("total_credit")),
            _num(record.get("bull_strangle_return_pct")),
            _num(record.get("call_distance_pct")),
            _num(record.get("put_distance_pct")),
            _num(record.get("buy_put_distance_pct")),
            _num(record.get("call_prob_otm")),
            _num(record.get("put_prob_otm")),
            _num(record.get("prob_both_otm")),
            record.get("selector_source"),
            _num(record.get("call_selector_pct")),
            _num(record.get("put_selector_pct")),
            _num(record.get("buy_put_selector_pct")),
            json.dumps(record, default=str, sort_keys=True),
        ),
    )


def _insert_deviation_row(conn, run_id: int, record: dict[str, Any], baseline: dict[str, Any]) -> None:
    stock_dev = _diff(record.get("live_stock_price"), baseline.get("stock_price"))
    raw = {
        "stock_price": {"live": record.get("live_stock_price"), "baseline": baseline.get("stock_price")},
        "iv": {"live": record.get("live_stock_iv"), "baseline": baseline.get("implied_volatility")},
        "sell_call_strike": {"live": record.get("sell_call_strike"), "baseline": baseline.get("sell_call_strike")},
        "sell_put_strike": {"live": record.get("sell_put_strike"), "baseline": baseline.get("sell_put_strike")},
        "buy_put_strike": {"live": record.get("buy_put_strike"), "baseline": baseline.get("buy_put_strike")},
        "total_credit": {
            "live": record.get("total_credit"),
            "baseline": _baseline_credit(baseline),
        },
    }
    conn.execute(
        """
        INSERT INTO watchlist_deviations
        (run_id, newsletter_id, newsletter_date, expiration_date, watchlist_entry_id, symbol,
         stock_price_deviation, stock_price_deviation_pct, iv_deviation,
         sell_call_strike_deviation, sell_put_strike_deviation, buy_put_strike_deviation,
         total_credit_deviation, raw_deviation_json)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            run_id,
            record.get("newsletter_id"),
            record.get("newsletter_date"),
            record.get("expiration_date"),
            record.get("watchlist_entry_id"),
            record.get("symbol"),
            stock_dev,
            _pct_diff(record.get("live_stock_price"), baseline.get("stock_price")),
            _diff(record.get("live_stock_iv"), baseline.get("implied_volatility")),
            _diff(record.get("sell_call_strike"), baseline.get("sell_call_strike")),
            _diff(record.get("sell_put_strike"), baseline.get("sell_put_strike")),
            _diff(record.get("buy_put_strike"), baseline.get("buy_put_strike")),
            _diff(record.get("total_credit"), _baseline_credit(baseline)),
            json.dumps(raw, default=str, sort_keys=True),
        ),
    )


def _baseline_credit(baseline: dict[str, Any]) -> float | None:
    values = [
        baseline.get("sell_call_premium"),
        baseline.get("sell_put_premium"),
        baseline.get("buy_put_premium"),
    ]
    if any(value is None for value in values):
        return None
    return float(values[0]) + float(values[1]) - float(values[2])


def _diff(live: Any, baseline: Any) -> float | None:
    live_num = _num(live)
    baseline_num = _num(baseline)
    if live_num is None or baseline_num is None:
        return None
    return live_num - baseline_num


def _pct_diff(live: Any, baseline: Any) -> float | None:
    live_num = _num(live)
    baseline_num = _num(baseline)
    if live_num is None or baseline_num in (None, 0):
        return None
    return (live_num - baseline_num) / baseline_num


def _string_or_none(value: Any) -> str | None:
    if value in (None, ""):
        return None
    return str(value)


def _int_or_none(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _num(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, str) and value.startswith("#"):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _sha256_file(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
