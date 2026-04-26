from __future__ import annotations

import json
import math
from hashlib import sha256
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .database import DEFAULT_DB_PATH, connect, initialize_database


DEFAULT_SELECTOR_ROUNDING_INCREMENT_PCT = 0.5
DEFAULT_TEMPLATE_VERSION = "os_live_v1"

DEFAULT_STRIKE_ROUNDING_POLICY = {
    "under_25": 0.5,
    "25_to_100": 1.0,
    "100_to_250": 2.5,
    "over_250": 5.0,
}

DELTA_FALLBACK_BANDS = {
    "days_to_expiration_min": 22,
    "days_to_expiration_max": 35,
    "buy_put_delta_min": 10.0,
    "buy_put_delta_max": 15.0,
    "sell_put_delta_min": 15.0,
    "sell_put_delta_max": 20.0,
    "sell_call_delta_min": 25.0,
    "sell_call_delta_max": 30.0,
}

OS_LIVE_HEADERS = [
    "newsletter_id",
    "newsletter_date",
    "expiration_date",
    "watchlist_entry_id",
    "symbol",
    "description",
    "baseline_stock_price",
    "baseline_iv",
    "baseline_sell_call_strike",
    "baseline_sell_put_strike",
    "baseline_buy_put_strike",
    "live_stock_price",
    "live_stock_iv",
    "live_sector",
    "live_industry",
    "live_earnings_date",
    "perf_1m",
    "perf_3m",
    "sma_50d",
    "sma_200d",
    "iv_rv_percentile",
    "atr_percent",
    "short_ratio",
    "sell_call_strike",
    "sell_call_bid",
    "sell_call_delta",
    "sell_put_strike",
    "sell_put_bid",
    "sell_put_delta",
    "buy_put_strike",
    "buy_put_ask",
    "buy_put_delta",
    "total_credit",
    "bull_strangle_return_pct",
    "call_distance_pct",
    "put_distance_pct",
    "buy_put_distance_pct",
    "call_prob_otm",
    "put_prob_otm",
    "prob_both_otm",
    "selector_source",
    "call_selector_pct",
    "put_selector_pct",
    "buy_put_selector_pct",
]

BASELINE_HEADERS = [
    "entry_id",
    "newsletter_id",
    "newsletter_date",
    "expiration_date",
    "symbol",
    "description",
    "sector",
    "stock_price",
    "implied_volatility",
    "total_open_interest",
    "industry",
    "sub_sector",
    "weekly_options",
    "latest_earnings",
    "sell_call_strike",
    "sell_call_premium",
    "sell_put_strike",
    "sell_put_premium",
    "buy_put_strike",
    "buy_put_premium",
    "bull_strangle_return_pct",
    "put_credit_spread_return_pct",
    "covered_call_return_pct",
    "is_favorite",
]


@dataclass(frozen=True)
class SelectorSet:
    newsletter_id: int
    newsletter_date: str
    expiration_date: str | None
    selector_source: str
    row_count: int
    valid_row_count: int
    avg_sell_call_distance_pct: float | None
    avg_sell_put_distance_pct: float | None
    avg_buy_put_distance_pct: float | None
    call_selector_pct: float | None
    put_selector_pct: float | None
    buy_put_selector_pct: float | None
    selector_rounding_increment_pct: float
    strike_rounding_policy: dict[str, float]

    def as_dict(self) -> dict[str, Any]:
        return {
            "newsletter_id": self.newsletter_id,
            "newsletter_date": self.newsletter_date,
            "expiration_date": self.expiration_date,
            "selector_source": self.selector_source,
            "row_count": self.row_count,
            "valid_row_count": self.valid_row_count,
            "avg_sell_call_distance_pct": self.avg_sell_call_distance_pct,
            "avg_sell_put_distance_pct": self.avg_sell_put_distance_pct,
            "avg_buy_put_distance_pct": self.avg_buy_put_distance_pct,
            "call_selector_pct": self.call_selector_pct,
            "put_selector_pct": self.put_selector_pct,
            "buy_put_selector_pct": self.buy_put_selector_pct,
            "selector_rounding_increment_pct": self.selector_rounding_increment_pct,
            "strike_rounding_policy": self.strike_rounding_policy,
        }


def round_to_increment(value: float | None, increment: float) -> float | None:
    if value is None:
        return None
    if increment <= 0:
        raise ValueError("Rounding increment must be greater than zero")
    scaled = value / increment
    if scaled >= 0:
        rounded_units = math.floor(scaled + 0.5)
    else:
        rounded_units = math.ceil(scaled - 0.5)
    return round(rounded_units * increment, 6)


def calculate_newsletter_selectors(
    newsletter_date: str,
    db_path: str | Path = DEFAULT_DB_PATH,
    selector_rounding_increment_pct: float = DEFAULT_SELECTOR_ROUNDING_INCREMENT_PCT,
) -> SelectorSet:
    initialize_database(db_path)
    with connect(db_path) as conn:
        newsletter = conn.execute(
            """
            SELECT newsletter_id, publication_date, target_expiration
            FROM newsletters
            WHERE publication_date = ?
            """,
            (newsletter_date,),
        ).fetchone()
        if not newsletter:
            raise ValueError(f"Newsletter not found for date: {newsletter_date}")

        stats = conn.execute(
            """
            SELECT
              COUNT(*) AS row_count,
              SUM(
                CASE
                  WHEN stock_price IS NOT NULL AND stock_price != 0
                   AND sell_call_strike IS NOT NULL
                   AND sell_put_strike IS NOT NULL
                   AND buy_put_strike IS NOT NULL
                  THEN 1 ELSE 0
                END
              ) AS valid_row_count,
              AVG(
                CASE
                  WHEN stock_price IS NOT NULL AND stock_price != 0
                   AND sell_call_strike IS NOT NULL
                  THEN (sell_call_strike - stock_price) / stock_price * 100
                END
              ) AS avg_sell_call_distance_pct,
              AVG(
                CASE
                  WHEN stock_price IS NOT NULL AND stock_price != 0
                   AND sell_put_strike IS NOT NULL
                  THEN (sell_put_strike - stock_price) / stock_price * 100
                END
              ) AS avg_sell_put_distance_pct,
              AVG(
                CASE
                  WHEN stock_price IS NOT NULL AND stock_price != 0
                   AND buy_put_strike IS NOT NULL
                  THEN (buy_put_strike - stock_price) / stock_price * 100
                END
              ) AS avg_buy_put_distance_pct
            FROM watchlist_entries
            WHERE newsletter_id = ?
            """,
            (newsletter["newsletter_id"],),
        ).fetchone()

    avg_call = _maybe_float(stats["avg_sell_call_distance_pct"])
    avg_sell_put = _maybe_float(stats["avg_sell_put_distance_pct"])
    avg_buy_put = _maybe_float(stats["avg_buy_put_distance_pct"])
    selector_source = (
        "newsletter_average"
        if avg_call is not None and avg_sell_put is not None and avg_buy_put is not None
        else "delta_fallback"
    )
    return SelectorSet(
        newsletter_id=int(newsletter["newsletter_id"]),
        newsletter_date=str(newsletter["publication_date"]),
        expiration_date=newsletter["target_expiration"],
        selector_source=selector_source,
        row_count=int(stats["row_count"] or 0),
        valid_row_count=int(stats["valid_row_count"] or 0),
        avg_sell_call_distance_pct=avg_call,
        avg_sell_put_distance_pct=avg_sell_put,
        avg_buy_put_distance_pct=avg_buy_put,
        call_selector_pct=round_to_increment(avg_call, selector_rounding_increment_pct),
        put_selector_pct=round_to_increment(avg_sell_put, selector_rounding_increment_pct),
        buy_put_selector_pct=round_to_increment(avg_buy_put, selector_rounding_increment_pct),
        selector_rounding_increment_pct=selector_rounding_increment_pct,
        strike_rounding_policy=dict(DEFAULT_STRIKE_ROUNDING_POLICY),
    )


def prepare_os_workbook_record(
    newsletter_date: str,
    db_path: str | Path = DEFAULT_DB_PATH,
    template_version: str = DEFAULT_TEMPLATE_VERSION,
    selector_rounding_increment_pct: float = DEFAULT_SELECTOR_ROUNDING_INCREMENT_PCT,
    generated_path: str | None = None,
    status: str = "planned",
) -> dict[str, Any]:
    selectors = calculate_newsletter_selectors(
        newsletter_date=newsletter_date,
        db_path=db_path,
        selector_rounding_increment_pct=selector_rounding_increment_pct,
    )
    formula_contract = build_formula_contract(selectors, template_version)
    with connect(db_path) as conn:
        conn.execute(
            """
            INSERT INTO os_workbooks
            (newsletter_id, newsletter_date, expiration_date, generated_path,
             template_version, selector_source,
             avg_sell_call_distance_pct, avg_sell_put_distance_pct, avg_buy_put_distance_pct,
             call_selector_pct, put_selector_pct, buy_put_selector_pct,
             selector_rounding_increment_pct, strike_rounding_policy_json,
             buy_put_delta_min, buy_put_delta_max,
             sell_put_delta_min, sell_put_delta_max,
             sell_call_delta_min, sell_call_delta_max,
             formula_contract_json, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(newsletter_id, template_version, selector_source) DO UPDATE SET
              expiration_date = excluded.expiration_date,
              generated_path = excluded.generated_path,
              avg_sell_call_distance_pct = excluded.avg_sell_call_distance_pct,
              avg_sell_put_distance_pct = excluded.avg_sell_put_distance_pct,
              avg_buy_put_distance_pct = excluded.avg_buy_put_distance_pct,
              call_selector_pct = excluded.call_selector_pct,
              put_selector_pct = excluded.put_selector_pct,
              buy_put_selector_pct = excluded.buy_put_selector_pct,
              selector_rounding_increment_pct = excluded.selector_rounding_increment_pct,
              strike_rounding_policy_json = excluded.strike_rounding_policy_json,
              buy_put_delta_min = excluded.buy_put_delta_min,
              buy_put_delta_max = excluded.buy_put_delta_max,
              sell_put_delta_min = excluded.sell_put_delta_min,
              sell_put_delta_max = excluded.sell_put_delta_max,
              sell_call_delta_min = excluded.sell_call_delta_min,
              sell_call_delta_max = excluded.sell_call_delta_max,
              formula_contract_json = excluded.formula_contract_json,
              status = excluded.status
            """,
            (
                selectors.newsletter_id,
                selectors.newsletter_date,
                selectors.expiration_date,
                generated_path,
                template_version,
                selectors.selector_source,
                selectors.avg_sell_call_distance_pct,
                selectors.avg_sell_put_distance_pct,
                selectors.avg_buy_put_distance_pct,
                selectors.call_selector_pct,
                selectors.put_selector_pct,
                selectors.buy_put_selector_pct,
                selectors.selector_rounding_increment_pct,
                json.dumps(selectors.strike_rounding_policy, sort_keys=True),
                DELTA_FALLBACK_BANDS["buy_put_delta_min"],
                DELTA_FALLBACK_BANDS["buy_put_delta_max"],
                DELTA_FALLBACK_BANDS["sell_put_delta_min"],
                DELTA_FALLBACK_BANDS["sell_put_delta_max"],
                DELTA_FALLBACK_BANDS["sell_call_delta_min"],
                DELTA_FALLBACK_BANDS["sell_call_delta_max"],
                json.dumps(formula_contract, sort_keys=True),
                status,
            ),
        )
        workbook = conn.execute(
            """
            SELECT *
            FROM os_workbooks
            WHERE newsletter_id = ? AND template_version = ? AND selector_source = ?
            """,
            (selectors.newsletter_id, template_version, selectors.selector_source),
        ).fetchone()
        conn.commit()
    result = dict(workbook)
    result["formula_contract_json"] = json.loads(result["formula_contract_json"])
    result["strike_rounding_policy_json"] = json.loads(result["strike_rounding_policy_json"])
    return result


def generate_os_workbook(
    newsletter_date: str,
    db_path: str | Path = DEFAULT_DB_PATH,
    output_dir: str | Path = "outputs/workbooks",
    template_version: str = DEFAULT_TEMPLATE_VERSION,
) -> dict[str, Any]:
    initialize_database(db_path)
    workbook_record = prepare_os_workbook_record(
        newsletter_date=newsletter_date,
        db_path=db_path,
        template_version=template_version,
        status="generating",
    )
    output_path = Path(output_dir) / f"BullStrangle_OS_Live_{newsletter_date}.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    rows = _load_watchlist_rows(db_path, int(workbook_record["newsletter_id"]))
    if not rows:
        raise ValueError(f"No watchlist rows found for newsletter date: {newsletter_date}")

    wb = Workbook()
    os_sheet = wb.active
    os_sheet.title = "OS_Live"
    baseline_sheet = wb.create_sheet("Baseline")
    instructions_sheet = wb.create_sheet("Instructions")
    metadata_sheet = wb.create_sheet("Metadata")
    metadata_sheet.sheet_state = "hidden"

    _build_os_live_sheet(os_sheet, rows, workbook_record)
    _build_baseline_sheet(baseline_sheet, rows)
    _build_instructions_sheet(instructions_sheet)
    _build_metadata_sheet(metadata_sheet, workbook_record)

    wb.save(output_path)
    workbook_hash = _sha256_file(output_path)
    abs_output_path = str(output_path.resolve())

    with connect(db_path) as conn:
        conn.execute(
            """
            UPDATE os_workbooks
            SET generated_path = ?, workbook_hash = ?, status = 'generated'
            WHERE workbook_id = ?
            """,
            (abs_output_path, workbook_hash, workbook_record["workbook_id"]),
        )
        conn.commit()

    workbook_record["generated_path"] = abs_output_path
    workbook_record["workbook_hash"] = workbook_hash
    workbook_record["status"] = "generated"
    workbook_record["row_count"] = len(rows)
    return workbook_record


def build_formula_contract(selectors: SelectorSet, template_version: str) -> dict[str, Any]:
    return {
        "template_version": template_version,
        "selector_source": selectors.selector_source,
        "expiration_cell": "$B$3",
        "call_selector_cell": "$B$4",
        "put_selector_cell": "$B$5",
        "buy_put_selector_cell": "$B$6",
        "selector_source_cell": "$B$7",
        "selector_rounding_increment_pct": selectors.selector_rounding_increment_pct,
        "call_selector_pct": selectors.call_selector_pct,
        "put_selector_pct": selectors.put_selector_pct,
        "buy_put_selector_pct": selectors.buy_put_selector_pct,
        "strike_rounding_policy": selectors.strike_rounding_policy,
        "delta_fallback_bands": DELTA_FALLBACK_BANDS,
        "option_formula_examples": {
            "sell_call_strike": '=_xldudf_optionsamurai_option(E10,"CALL",$B$3,$B$4,"strike")',
            "sell_call_bid": '=_xldudf_optionsamurai_option(E10,"CALL",$B$3,$B$4,"bid")',
            "sell_put_strike": '=_xldudf_optionsamurai_option(E10,"PUT",$B$3,$B$5,"strike")',
            "sell_put_bid": '=_xldudf_optionsamurai_option(E10,"PUT",$B$3,$B$5,"bid")',
        },
    }


def _maybe_float(value: Any) -> float | None:
    if value is None:
        return None
    return float(value)


def _load_watchlist_rows(db_path: str | Path, newsletter_id: int) -> list[dict[str, Any]]:
    with connect(db_path) as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM watchlist_entries
            WHERE newsletter_id = ?
            ORDER BY symbol
            """,
            (newsletter_id,),
        ).fetchall()
    return [dict(row) for row in rows]


def _build_os_live_sheet(ws, rows: list[dict[str, Any]], workbook_record: dict[str, Any]) -> None:
    ws.freeze_panes = "A10"
    ws["A1"] = "BullStrangle OS Live Workbook"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    params = [
        ("Newsletter Date", workbook_record["newsletter_date"]),
        ("Expiration Date", workbook_record["expiration_date"]),
        ("Call Selector", _format_selector(workbook_record["call_selector_pct"])),
        ("Put Selector", _format_selector(workbook_record["put_selector_pct"])),
        ("Buy Put Selector", _format_selector(workbook_record["buy_put_selector_pct"])),
        ("Selector Source", workbook_record["selector_source"]),
    ]
    for idx, (label, value) in enumerate(params, start=2):
        ws.cell(row=idx, column=1, value=label)
        ws.cell(row=idx, column=2, value=value)
        ws.cell(row=idx, column=1).font = Font(bold=True)

    header_row = 9
    for col_idx, header in enumerate(OS_LIVE_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="5B9BD5")
        cell.alignment = Alignment(horizontal="center")

    for row_idx, entry in enumerate(rows, start=10):
        _write_os_live_row(ws, row_idx, entry, workbook_record)

    _format_sheet(ws, widths={
        "A": 14, "B": 14, "C": 14, "D": 18, "E": 12, "F": 26,
        "L": 16, "M": 14, "X": 16, "Y": 14, "AA": 16, "AB": 14,
        "AD": 16, "AE": 14, "AG": 14, "AH": 18,
    })
    ws.auto_filter.ref = f"A9:{get_column_letter(len(OS_LIVE_HEADERS))}{len(rows) + 9}"


def _write_os_live_row(ws, row_idx: int, entry: dict[str, Any], workbook_record: dict[str, Any]) -> None:
    symbol_cell = f"E{row_idx}"
    expiration_ref = "$B$3"
    call_selector_ref = "$B$4"
    put_selector_ref = "$B$5"
    buy_put_selector_ref = "$B$6"

    values: list[Any] = [
        entry["newsletter_id"],
        entry["newsletter_date"],
        entry["expiration_date"],
        entry["entry_id"],
        entry["symbol"],
        entry["description"],
        entry["stock_price"],
        entry["implied_volatility"],
        entry["sell_call_strike"],
        entry["sell_put_strike"],
        entry["buy_put_strike"],
        f'=_xldudf_optionsamurai_stock({symbol_cell},"stock_last")',
        f'=_xldudf_optionsamurai_stock({symbol_cell},"stock_iv")',
        f'=_xldudf_optionsamurai_stock({symbol_cell},"sector")',
        f'=_xldudf_optionsamurai_stock({symbol_cell},"industry")',
        f'=_xldudf_optionsamurai_stock({symbol_cell},"earnings_date")',
        f'=_xldudf_optionsamurai_stock({symbol_cell},"perf_m")',
        f'=_xldudf_optionsamurai_stock({symbol_cell},"perf_q")',
        f'=_xldudf_optionsamurai_stock({symbol_cell},"sma_50d")',
        f'=_xldudf_optionsamurai_stock({symbol_cell},"sma_200d")',
        f'=_xldudf_optionsamurai_stock({symbol_cell},"stock_iv_rv_pr")',
        f'=_xldudf_optionsamurai_stock({symbol_cell},"atr_percent")',
        f'=_xldudf_optionsamurai_stock({symbol_cell},"short_ratio")',
        f'=_xldudf_optionsamurai_option({symbol_cell},"CALL",{expiration_ref},{call_selector_ref},"strike")',
        f'=_xldudf_optionsamurai_option({symbol_cell},"CALL",{expiration_ref},{call_selector_ref},"bid")',
        f'=_xldudf_optionsamurai_option({symbol_cell},"CALL",{expiration_ref},{call_selector_ref},"delta")',
        f'=_xldudf_optionsamurai_option({symbol_cell},"PUT",{expiration_ref},{put_selector_ref},"strike")',
        f'=_xldudf_optionsamurai_option({symbol_cell},"PUT",{expiration_ref},{put_selector_ref},"bid")',
        f'=_xldudf_optionsamurai_option({symbol_cell},"PUT",{expiration_ref},{put_selector_ref},"delta")',
        f'=_xldudf_optionsamurai_option({symbol_cell},"PUT",{expiration_ref},{buy_put_selector_ref},"strike")',
        f'=_xldudf_optionsamurai_option({symbol_cell},"PUT",{expiration_ref},{buy_put_selector_ref},"ask")',
        f'=_xldudf_optionsamurai_option({symbol_cell},"PUT",{expiration_ref},{buy_put_selector_ref},"delta")',
        f"=Y{row_idx}+AB{row_idx}-AE{row_idx}",
        f'=IF(G{row_idx}>0,AG{row_idx}/G{row_idx},"")',
        f'=IF(L{row_idx}>0,(X{row_idx}-L{row_idx})/L{row_idx},"")',
        f'=IF(L{row_idx}>0,(AA{row_idx}-L{row_idx})/L{row_idx},"")',
        f'=IF(L{row_idx}>0,(AD{row_idx}-L{row_idx})/L{row_idx},"")',
        f'=_xldudf_optionsamurai_option({symbol_cell},"CALL",{expiration_ref},TEXT(X{row_idx},"0"),"prob_otm")',
        f'=_xldudf_optionsamurai_option({symbol_cell},"PUT",{expiration_ref},TEXT(AA{row_idx},"0"),"prob_otm")',
        f'=IF(AND(AL{row_idx}<>"",AM{row_idx}<>""),AL{row_idx}*AM{row_idx},"")',
        workbook_record["selector_source"],
        workbook_record["call_selector_pct"],
        workbook_record["put_selector_pct"],
        workbook_record["buy_put_selector_pct"],
    ]
    for col_idx, value in enumerate(values, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)


def _build_baseline_sheet(ws, rows: list[dict[str, Any]]) -> None:
    ws.freeze_panes = "A2"
    for col_idx, header in enumerate(BASELINE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="70AD47")
    for row_idx, entry in enumerate(rows, start=2):
        for col_idx, header in enumerate(BASELINE_HEADERS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=entry.get(header))
    _format_sheet(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(BASELINE_HEADERS))}{len(rows) + 1}"


def _build_instructions_sheet(ws) -> None:
    rows = [
        ["BullStrangle OS Workbook"],
        ["1. Open this workbook with the Option Samurai Excel add-in enabled."],
        ["2. Refresh formulas during market hours."],
        ["3. Save the workbook after formulas finish refreshing."],
        ["4. Upload the refreshed workbook back to MCP for daily ingestion."],
        ["5. Final DCA and Bull Strangle decisions are generated only after the week ends."],
    ]
    for row_idx, row in enumerate(rows, start=1):
        ws.cell(row=row_idx, column=1, value=row[0])
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws.column_dimensions["A"].width = 110


def _build_metadata_sheet(ws, workbook_record: dict[str, Any]) -> None:
    metadata = {
        key: value
        for key, value in workbook_record.items()
        if key not in {"formula_contract_json", "strike_rounding_policy_json"}
    }
    metadata["formula_contract_json"] = json.dumps(
        workbook_record["formula_contract_json"], sort_keys=True
    )
    metadata["strike_rounding_policy_json"] = json.dumps(
        workbook_record["strike_rounding_policy_json"], sort_keys=True
    )
    ws.append(["key", "value"])
    for key, value in metadata.items():
        ws.append([key, value])
    _format_sheet(ws)


def _format_sheet(ws, widths: dict[str, int] | None = None) -> None:
    widths = widths or {}
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = widths.get(letter, 16)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, float):
                cell.number_format = "0.00"


def _format_selector(value: float | None) -> str | None:
    if value is None:
        return None
    return f"{value:.1f}%"


def _sha256_file(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
