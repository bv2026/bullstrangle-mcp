import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from bullstrangle_mcp.decisions import generate_weekend_decisions
from bullstrangle_mcp.ingestion import ingest_newsletter
from bullstrangle_mcp.os_ingestion import ingest_os_workbook
from bullstrangle_mcp.os_reports import report_os_run
from bullstrangle_mcp.os_weekly import aggregate_os_week
from bullstrangle_mcp.os_workbooks import (
    calculate_newsletter_selectors,
    generate_os_workbook,
    prepare_os_workbook_record,
)
from bullstrangle_mcp.positions import ingest_positions
from bullstrangle_mcp.database import connect
from bullstrangle_mcp.tools import get_newsletter_by_ref_tool, get_newsletter_tool


SAMPLE_PDF = Path("data/newsletters/Bull-Strangle-Weekly-Newsletter-For-Week-End-Apr-17-2026.pdf")


def require_sample_pdf() -> Path:
    if not SAMPLE_PDF.exists():
        pytest.skip(f"Sample newsletter PDF not available: {SAMPLE_PDF}")
    return SAMPLE_PDF


@pytest.mark.integration
def test_ingest_sample_newsletter(tmp_path):
    pdf = require_sample_pdf()
    db = tmp_path / "bullstrangle.db"

    result = ingest_newsletter(pdf, db)

    assert result["publication_date"] == "2026-04-17"
    assert result["target_expiration"] == "2026-05-15"
    assert result["watchlist_count"] >= 20
    assert result["hybrid_score"] == 2

    stored = get_newsletter_tool(result["newsletter_id"], str(db))
    assert stored["market_environment"]["sp500_price"] == 7126.06
    assert any(row["symbol"] == "NTAP" for row in stored["watchlist"])

    stored_by_date = get_newsletter_by_ref_tool("2026-04-17", str(db))
    assert stored_by_date["newsletter"]["newsletter_id"] == result["newsletter_id"]
    assert any(row["symbol"] == "NTAP" for row in stored_by_date["watchlist"])


@pytest.mark.integration
def test_prepare_os_workbook_selectors(tmp_path):
    pdf = require_sample_pdf()
    db = tmp_path / "bullstrangle.db"
    ingest_newsletter(pdf, db)

    selectors = calculate_newsletter_selectors("2026-04-17", db)

    assert selectors.selector_source == "newsletter_average"
    assert selectors.call_selector_pct == 4.0
    assert selectors.put_selector_pct == -3.5
    assert selectors.buy_put_selector_pct == -13.5

    workbook = prepare_os_workbook_record("2026-04-17", db)

    assert workbook["newsletter_date"] == "2026-04-17"
    assert workbook["selector_source"] == "newsletter_average"
    assert workbook["call_selector_pct"] == 4.0
    assert workbook["put_selector_pct"] == -3.5
    assert workbook["formula_contract_json"]["selector_rounding_increment_pct"] == 0.5


@pytest.mark.integration
def test_generate_os_workbook(tmp_path):
    pdf = require_sample_pdf()
    db = tmp_path / "bullstrangle.db"
    output_dir = tmp_path / "os_workbooks"
    ingest_newsletter(pdf, db)

    result = generate_os_workbook("2026-04-17", db, output_dir)

    output_path = Path(result["generated_path"])
    assert output_path.exists()
    assert result["status"] == "generated"
    assert result["workbook_hash"]

    workbook = load_workbook(output_path, data_only=False)
    assert workbook.sheetnames[:3] == ["OS_Live", "Baseline", "Instructions"]
    assert workbook["OS_Live"]["B3"].value == "2026-05-15"
    assert workbook["OS_Live"]["B4"].value == "4.0%"
    assert workbook["OS_Live"]["B5"].value == "-3.5%"
    assert workbook["OS_Live"]["L10"].value == '=_xldudf_optionsamurai_stock(E10,"stock_last")'
    assert workbook["OS_Live"]["X10"].value == '=_xldudf_optionsamurai_option(E10,"CALL",$B$3,$B$4,"strike")'


@pytest.mark.integration
def test_ingest_os_workbook(tmp_path):
    pdf = require_sample_pdf()
    db = tmp_path / "bullstrangle.db"
    output_dir = tmp_path / "os_workbooks"
    ingest_newsletter(pdf, db)
    generated = generate_os_workbook("2026-04-17", db, output_dir)

    result = ingest_os_workbook(generated["generated_path"], db, trading_date="2026-04-22")

    assert result["newsletter_date"] == "2026-04-17"
    assert result["row_count"] == 24
    assert result["formula_cell_count"] > 0
    assert result["status"] in {"ingested", "ingested_no_cached_values"}

    report = report_os_run(result["run_id"], db)

    assert report["run"]["run_id"] == result["run_id"]
    assert report["counts"]["evaluation_rows"] == 24
    assert "OS Run Report" in report["markdown"]

    aggregate = aggregate_os_week("2026-04-17", db)

    assert aggregate["newsletter_date"] == "2026-04-17"
    assert aggregate["run_count"] == 1
    assert aggregate["symbol_count"] == 24
    assert aggregate["valid_symbol_count"] <= 24
    assert "Weekly OS Aggregation" in aggregate["markdown"]

    output_path = tmp_path / "weekly_aggregate.md"
    written = aggregate_os_week("2026-04-17", db, output_path)

    assert Path(written["output_path"]) == output_path
    assert output_path.exists()
    assert "Weekly OS Aggregation" in output_path.read_text(encoding="utf-8")

    positions_csv = tmp_path / "positions.csv"
    positions_csv.write_text(
        "\n".join(
            [
                "ACCOUNT,Symbol,Quantity, Price , AVG PRICE , Market Value , Cost Basis ",
                'A1,IBIT,50, $52.00 , $50.00 ," $2600.00 "," $2500.00 "',
                'A1,ZIM,100, $17.00 , $16.00 ," $1700.00 "," $1600.00 "',
            ]
        ),
        encoding="utf-8",
    )
    positions = ingest_positions(positions_csv, db)

    decisions = generate_weekend_decisions(
        "2026-04-17",
        db,
        decision_date="2026-04-25",
        output_path=tmp_path / "weekend_decisions.md",
    )

    assert decisions["newsletter_date"] == "2026-04-17"
    assert decisions["os_run_count"] == 1
    assert decisions["position_run_id"] == positions["position_run_id"]
    assert decisions["positions_available"] is True
    assert sum(decisions["selected_action_counts"].values()) == 24
    assert decisions["bull_strangle_counts"]["SKIP"] >= 1
    assert "Weekend Decisions" in decisions["markdown"]
    assert (tmp_path / "weekend_decisions.md").exists()

    with connect(db) as conn:
        batch = conn.execute(
            """
            SELECT position_run_id, strategy_logic_version
            FROM decision_batches
            WHERE newsletter_date = ? AND decision_date = ?
            """,
            ("2026-04-17", "2026-04-25"),
        ).fetchone()
        assert batch["position_run_id"] == positions["position_run_id"]
        assert batch["strategy_logic_version"] == "score_branch_v1"
        bull_row = conn.execute(
            """
            SELECT selected_action, strategy_score, strategy_band,
                   rules_passed_json, rules_failed_json, source_snapshot_json
            FROM bull_strangle_decisions
            WHERE decision_batch_id = ? AND symbol = 'ZIM'
            """,
            (decisions["decision_batch_id"],),
        ).fetchone()
        assert bull_row["selected_action"] in {"BULL_STRANGLE", "WATCH", "SKIP"}
        assert bull_row["strategy_score"] is not None
        assert bull_row["strategy_band"] in {"strong", "moderate", "watch", "weak"}
        passed = json.loads(bull_row["rules_passed_json"])
        failed = json.loads(bull_row["rules_failed_json"])
        snapshot = json.loads(bull_row["source_snapshot_json"])
        assert passed["selected_action"] in {"BULL_STRANGLE", "DCA", "WATCH", "SKIP"}
        assert isinstance(failed, dict)
        assert snapshot["strategy_context"]["strategy_score"] is not None


@pytest.mark.integration
def test_generate_weekend_decisions_rule_sensitivity_via_db(tmp_path):
    """DB-modified thresholds are reflected in stored decisions; DCA path completes cleanly.

    Two things are verified here:

    1. Rules loaded from strategy_rules appear in rules_applied_json — confirming
       load_decision_rules() is actually wired into the pipeline (not bypassed).

    2. The DCA decision path completes without KeyError when rules are non-default.
       This is the regression test for the bug caught in the 2026-04-23 e2e review:
       _build_dca_decision was slicing rules to the DCA sub-dict and forwarding
       that slice to _build_rule_diagnostics, which needs bull_strangle.max_credit_deviation.
    """
    pdf = require_sample_pdf()
    db = tmp_path / "bullstrangle.db"
    output_dir = tmp_path / "os_workbooks"

    ingest_newsletter(pdf, db)
    generated = generate_os_workbook("2026-04-17", db, output_dir)
    ingest_os_workbook(generated["generated_path"], db, trading_date="2026-04-22")

    positions_csv = tmp_path / "positions.csv"
    positions_csv.write_text(
        "\n".join([
            "ACCOUNT,Symbol,Quantity, Price , AVG PRICE , Market Value , Cost Basis ",
            'A1,IBIT,50, $52.00 , $50.00 ," $2600.00 "," $2500.00 "',
            'A1,ZIM,100, $17.00 , $16.00 ," $1700.00 "," $1600.00 "',
        ]),
        encoding="utf-8",
    )
    ingest_positions(positions_csv, db)

    # Run once with default (seeded) rules so a baseline batch exists.
    decisions_default = generate_weekend_decisions("2026-04-17", db, decision_date="2026-04-25")
    assert sum(decisions_default["bull_strangle_counts"].values()) == 24
    assert sum(decisions_default["dca_counts"].values()) == 24

    # Tighten max_credit_deviation to a non-default value in the DB.
    custom_value = 0.50  # default is 2.50
    with connect(db) as conn:
        conn.execute(
            "UPDATE strategy_rules SET rule_parameters = ? WHERE rule_name = ?",
            (f'{{"value":{custom_value}}}', "bull_strangle_max_credit_deviation"),
        )

    # Re-run: must complete without KeyError (DCA path regression).
    decisions_tight = generate_weekend_decisions("2026-04-17", db, decision_date="2026-04-26")

    # All 24 symbols processed for both decision types.
    assert sum(decisions_tight["bull_strangle_counts"].values()) == 24
    assert sum(decisions_tight["dca_counts"].values()) == 24  # KeyError would abort this

    # Verify the DB-modified value is stored in rules_applied_json — confirms
    # load_decision_rules() is wired into the pipeline and not bypassed.
    with connect(db) as conn:
        sample = conn.execute(
            """
            SELECT rules_applied_json
            FROM bull_strangle_decisions
            WHERE decision_batch_id = ?
            LIMIT 1
            """,
            (decisions_tight["decision_batch_id"],),
        ).fetchone()
    applied = json.loads(sample["rules_applied_json"])
    assert applied["max_credit_deviation"] == pytest.approx(custom_value), (
        f"Expected rules_applied_json to reflect DB value {custom_value}, got {applied}"
    )
